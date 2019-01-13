from openpyxl import load_workbook

excel_oku = load_workbook(filename= "8-A-RAPOR.xlsx")
excel_yaz = load_workbook(filename= "karne.xlsx")
sinifSayisi=0
hucreNo_oku = 5
hucreNo_yaz = 17


while hucreNo_oku<21:

    sayfa_oku = excel_oku[excel_oku.sheetnames[0]]
    sayfa_yaz = excel_yaz[excel_yaz.sheetnames[sinifSayisi]]

    while sayfa_yaz["C" + str(hucreNo_yaz)].value != None:  # hücre boş olmadığı sürece diğer satıra geç!
        hucreNo_yaz += 1


    def oku(harf,rakam):
        return sayfa_oku[harf+str(rakam)].value


    def yaz(harf,rakam,deger):
       sayfa_yaz[harf+str(rakam)] = deger

    yaz("C", hucreNo_yaz, deger=oku("G", 2))            # Tarih
    yaz("G", hucreNo_yaz, deger=oku("H", hucreNo_oku))  # Türkçe
    yaz("H",hucreNo_yaz, deger = oku("I",hucreNo_oku))  #Türkçe
    yaz("I",hucreNo_yaz, deger = oku("J",hucreNo_oku))  #Türkçe
    yaz("J", hucreNo_yaz, deger=oku("K", hucreNo_oku))  # Türkçe
    yaz("K", hucreNo_yaz, deger=oku("L", hucreNo_oku))  # Türkçe
    yaz("N",hucreNo_yaz, deger = oku("O",hucreNo_oku)) #Matematik
    yaz("O",hucreNo_yaz, deger = oku("P",hucreNo_oku))  #Matematik
    yaz("P",hucreNo_yaz, deger = oku("Q",hucreNo_oku)) #Matematik
    yaz("Q", hucreNo_yaz, deger=oku("R", hucreNo_oku))  # Matematik
    yaz("R", hucreNo_yaz, deger=oku("S", hucreNo_oku))  # Matematik
    yaz("S", hucreNo_yaz, deger=oku("T", hucreNo_oku))  # Matematik
    yaz("V", hucreNo_yaz, deger=oku("W", hucreNo_oku))  # Fen
    yaz("W", hucreNo_yaz, deger=oku("X", hucreNo_oku))  # Fen
    yaz("X", hucreNo_yaz, deger=oku("Y", hucreNo_oku))  # Fen
    yaz("Y", hucreNo_yaz, deger=oku("Z", hucreNo_oku))  # Fen
    yaz("Z", hucreNo_yaz, deger=oku("AA", hucreNo_oku))  # Fen
    yaz("AC", hucreNo_yaz, deger=oku("AD", hucreNo_oku))  # Sosyal
    yaz("AD", hucreNo_yaz, deger=oku("AE", hucreNo_oku))  # Sosyal
    yaz("AE", hucreNo_yaz, deger=oku("AF", hucreNo_oku))  # Sosyal
    yaz("AF", hucreNo_yaz, deger=oku("AG", hucreNo_oku))  # Sosyal
    yaz("AG", hucreNo_yaz, deger=oku("AH", hucreNo_oku))  # Sosyal
    yaz("AJ", hucreNo_yaz, deger=oku("AK", hucreNo_oku))  # İngilizce
    yaz("AK", hucreNo_yaz, deger=oku("AL", hucreNo_oku))  # İngilizce
    yaz("AL", hucreNo_yaz, deger=oku("AM", hucreNo_oku))  # İngilizce
    yaz("AM", hucreNo_yaz, deger=oku("AN", hucreNo_oku))  # İngilizce
    yaz("AN", hucreNo_yaz, deger=oku("AO", hucreNo_oku))  # İngilizce
    yaz("AQ", hucreNo_yaz, deger=oku("AR", hucreNo_oku))  # Din
    hucreNo_oku+=1
    sinifSayisi+=1
print("yazma işlemi bitti")
print(hucreNo_oku)

excel_yaz.save("karne.xlsx")
excel_yaz.close()